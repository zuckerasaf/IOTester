import math
import subprocess
from pathlib import Path
import textwrap
import openpyxl
import pandas as pd

BASE_DIR = Path(__file__).resolve().parent


file_name = "IO_Tester_logic_Power_test"
XLSM_PATH = BASE_DIR / f"{file_name}.xlsm"

OUT_DOT = BASE_DIR / f"{file_name}.dot"
OUT_SVG = BASE_DIR / f"{file_name}.svg"
OUT_HTML = BASE_DIR / f"{file_name}.html"

SHEET = "Sheet1"


def wrap(s, width=28):
    s = fmt(s)
    if not s:
        return ""
    return "\\n".join(textwrap.wrap(s, width=width))


def fmt(v):
    """Convert Excel cell value to clean printable text."""
    if v is None or is_nan(v):
        return ""
    s = str(v).replace("\r", " ").replace("\n", " ").strip()
    return s

def gv_escape(s: str) -> str:
    """Escape text for Graphviz quoted labels."""
    return s.replace('"', r'\"')


def is_nan(x):
    return isinstance(x, float) and math.isnan(x)

wb = openpyxl.load_workbook(XLSM_PATH, keep_vba=True, data_only=False)
ws = wb[SHEET]

headers = [ws.cell(1, c).value for c in range(1, 9)]  # A..H
rows = []
for r in range(2, ws.max_row + 1):
    key = ws.cell(r, 1).value
    if key is None:
        continue
    row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, 9)}
    rows.append(row)

df = pd.DataFrame(rows)
df = df[df[headers[0]].notna()]  # key exists

def safe_label(v):
    if v is None:
        return ""
    return str(v).replace('"', r'\"').replace("\r", "").strip()

# Build DOT
dot = []
dot.append('digraph G {')
dot.append('rankdir=TB;')
dot.append(
  'node [shape=box, style="rounded,filled", fillcolor="#F7F7F7", color="#666666", '
  'fontname="Arial", fontsize=11, fixedsize=false, width=3.2];'
)
dot.append('edge [color="#999999"];')

# Example: A=key, B=name, C=parent, D=step_type, E=expected, F=measured
COL_KEY = 0
COL_NAME = 1
COL_PARENT = 2
COL_DEFINITION = 4
COL_COMMENT = 5
COL_DOS = 6
COL_DOD = 7

for _, row in df.iterrows():
    

    key = int(row[headers[COL_KEY]])
    name = wrap(row[headers[COL_NAME]], width=28)
    definition = wrap(row[headers[COL_DEFINITION]], width=28)
    comment = wrap(row[headers[COL_COMMENT]], width=28)
    DOS = wrap(row[headers[COL_DOS]], width=28)
    DOD = wrap(row[headers[COL_DOD]], width=28)

    # name = fmt(row[headers[COL_NAME]])
    # definition = fmt(row[headers[COL_DEFINITION]])
    # comment = fmt(row[headers[COL_COMMENT]])
    # DOS = fmt(row[headers[COL_DOS]])
    # DOD = fmt(row[headers[COL_DOD]])


    # Build multi-line label (Graphviz uses \n)
    lines = [
        name or f"Node {key}",
        f"Key: {key}",
    ]

    if definition:
        lines.append(f"Type: {definition}")
    if comment:
        lines.append(f"Expected: {comment}")
    if DOS:
        lines.append(f"DOS: {DOS}")
    if DOD:
        lines.append(f"DOD: {DOD}")

    label = gv_escape("\\n".join(lines))
    dot.append(f'{key} [id="n{key}", label="{label}"];')


for _, row in df.iterrows():
    key = int(row[headers[0]])
    parent = row[headers[2]]
    if parent is None or is_nan(parent):
        continue
    dot.append(f'{int(parent)} -> {key};')


dot.append('}')
OUT_DOT.write_text("\n".join(dot), encoding="utf-8")

# DOT -> SVG
subprocess.run(["dot", "-Tsvg", str(OUT_DOT), "-o", str(OUT_SVG)], check=True)

svg_content = OUT_SVG.read_text(encoding="utf-8")

html_doc = f"""<!doctype html>
<html>
<head>
<meta charset="utf-8"/>
<title>Logic Tree Viewer</title>
<style>
  body {{ font-family: Arial, sans-serif; margin: 0; }}
  header {{ padding: 10px 14px; background: #111827; color: #fff; display:flex; gap:12px; align-items:center; }}
  header input {{ width: 110px; padding:6px 8px; border-radius:8px; border:1px solid #374151; }}
  header button {{ padding:6px 10px; border-radius:8px; border:1px solid #374151; background:#1f2937; color:#fff; cursor:pointer; }}
  #wrap {{ padding: 10px; overflow: auto; height: calc(100vh - 52px); background: #f3f4f6; }}
  svg {{ background: white; border-radius: 12px; box-shadow: 0 8px 20px rgba(0,0,0,.08); }}
  g.node.active path {{ stroke-width: 4px; }}
  g.node.active text {{ font-weight: 700; }}
</style>
</head>
<body>
<header>
  <strong>IO Tester Logic Tree</strong>
  <span style="opacity:.8;">Key:</span>
  <input id="nodeKey" placeholder="450" />
  <button onclick="highlight(document.getElementById('nodeKey').value)">Highlight</button>
  <button onclick="clearAll()">Clear</button>
  <button onclick="manualRefresh()" style="background:#059669;">Force Refresh Trace</button>
  <button onclick="testUpdate()" style="background:#dc2626;">Test Update</button>
  <span id="statusInfo" style="margin-left:auto; opacity:.8; font-size:13px;">Initializing...</span>
</header>

<div id="wrap">
{svg_content}
</div>

<script>
  function findNodeGroupByKey(key) {{
    key = String(key).trim();
    if (!key) return null;

    // Debug: Log available nodes (only first time or when key is unusual)
    if (!window.nodesLogged) {{
      const allNodes = document.querySelectorAll('g.node title');
      const nodeKeys = Array.from(allNodes).map(t => t.textContent.trim());
      console.log('Available nodes in SVG:', nodeKeys.join(', '));
      window.nodesLogged = true;
    }}

    // Graphviz typically uses: <g class="node"> ... <title>450</title> ...
    const nodes = document.querySelectorAll('g.node');
    for (const g of nodes) {{
      const t = g.querySelector('title');
      if (t && t.textContent.trim() === key) {{
        console.log('✓ Found node for key:', key);
        return g;
      }}
    }}
    console.warn('✗ Node not found for key:', key);
    return null;
  }}

  function clearAll() {{
    document.querySelectorAll('g.node').forEach(g => g.classList.remove('active','done','fail'));
  }}

  function setNodeStatus(key, status) {{
    const g = findNodeGroupByKey(key);
    if (!g) return false;

    g.classList.remove('active','done','fail');
    g.classList.add(status);

    const shape = g.querySelector('polygon, path, ellipse');
    if (shape) {{
      if (status === 'active') shape.setAttribute('fill', '#FFF3B0');
      if (status === 'done')   shape.setAttribute('fill', '#D1FAE5');
      if (status === 'fail')   shape.setAttribute('fill', '#FECACA');
    }}
    return true;
  }}

  function highlight(key) {{
    clearAll();
    setNodeStatus(key, 'active');
    const g = findNodeGroupByKey(key);
    if (g) g.scrollIntoView({{behavior:'smooth', block:'center', inline:'center'}});
  }}

  window.traceStep = function(key, status) {{
    setNodeStatus(key, status || 'active');
  }}

// Server-Sent Events (SSE) for real-time updates
let eventSource = null;
let lastTs = 0;
let lastKey = null;
let lastStatus = null;
let updateCount = 0;
let reconnectAttempts = 0;
let maxReconnectAttempts = 5;

// Update status display
function updateStatusDisplay() {{
  const info = document.getElementById('statusInfo');
  if (info) {{
    const status = eventSource && eventSource.readyState === EventSource.OPEN ? '✓ Connected' : '✗ Disconnected';
    info.textContent = `${{status}} | Updates: ${{updateCount}} | Last: ${{lastKey || 'none'}}`;
  }}
}}

function connectSSE() {{
  console.log('[SSE] Connecting to event stream...');
  
  eventSource = new EventSource('/events');
  
  eventSource.onopen = function() {{
    console.log('✓ SSE connection established');
    reconnectAttempts = 0;
    updateStatusDisplay();
    // Immediately fetch current state
    fetchTraceData();
  }};
  
  eventSource.onmessage = function(event) {{
    console.log('[SSE] Received:', event.data);
    if (event.data === 'update') {{
      // Server notified us of an update - fetch trace.json
      fetchTraceData();
    }}
  }};
  
  eventSource.onerror = function(err) {{
    console.error('[SSE] Connection error:', err);
    eventSource.close();
    updateStatusDisplay();
    
    // Attempt to reconnect with exponential backoff
    if (reconnectAttempts < maxReconnectAttempts) {{
      reconnectAttempts++;
      const delay = Math.min(1000 * Math.pow(2, reconnectAttempts), 30000);
      console.log(`[SSE] Reconnecting in ${{delay}}ms (attempt ${{reconnectAttempts}}/${{maxReconnectAttempts}})...`);
      setTimeout(connectSSE, delay);
    }} else {{
      console.error('[SSE] Max reconnect attempts reached. Please refresh page.');
      document.getElementById('statusInfo').textContent = '✗ Connection lost - Refresh page';
    }}
  }};
}}

async function fetchTraceData() {{
  try {{
    // Simple fetch without aggressive caching (SSE tells us when to read)
    const res = await fetch('trace.json?t=' + Date.now(), {{
      cache: 'no-store'
    }});
    
    if (!res.ok) {{
      console.warn('[TRACE] File not found yet');
      return;
    }}

    const data = await res.json();
    
    // Validate data structure
    if (!data || typeof data.key === 'undefined') {{
      console.warn('[TRACE] Invalid data:', data);
      return;
    }}
    
    // Detect changes: timestamp OR key OR status changed
    const hasChanged = (
      data.ts !== lastTs || 
      data.key !== lastKey || 
      data.status !== lastStatus
    );
    
    if (!hasChanged) {{
      return; // No changes detected
    }}

    // Update detected!
    updateCount++;
    updateStatusDisplay();
    console.log('✓ Update #' + updateCount + ': Key=' + data.key + ', Status=' + data.status + ', TS=' + data.ts.toFixed(2));
    
    // Clear previous node if key changed
    if (lastKey !== null && lastKey !== data.key) {{
      setNodeStatus(lastKey, 'done');
      console.log('  → Marked previous node ' + lastKey + ' as done');
    }}
    
    // Update tracking variables
    lastTs = data.ts;
    lastKey = data.key;
    lastStatus = data.status;
    
    // Apply the new status
    const success = setNodeStatus(data.key, data.status);
    if (success) {{
      console.log('  ✓ Node ' + data.key + ' highlighted as "' + data.status + '"');
      
      // Scroll to the active node
      const g = findNodeGroupByKey(data.key);
      if (g) {{
        g.scrollIntoView({{behavior:'smooth', block:'center', inline:'center'}});
        console.log('  → Scrolled to node');
      }}
    }} else {{
      console.warn('  ✗ Node ' + data.key + ' NOT FOUND in SVG!');
    }}
  }} catch (e) {{
    console.error('[TRACE] Fetch error:', e.message);
  }}
}}

// Connect to SSE on page load
connectSSE();

// Fallback: also check once on page load
fetchTraceData();

// Manual controls for debugging
window.reconnectSSE = function() {{
  if (eventSource) {{
    eventSource.close();
  }}
  reconnectAttempts = 0;
  connectSSE();
}};

window.manualRefresh = function() {{
  lastTs = 0;
  lastKey = null;
  lastStatus = null;
  updateCount = 0;
  clearAll();
  updateStatusDisplay();
  console.log('✓ Manual refresh - fetching current state');
  fetchTraceData();
}};

window.checkStatus = function() {{
  console.log('───────────────────────────────');
  console.log('SSE State:');
  console.log('  ReadyState:', eventSource ? eventSource.readyState : 'null');
  console.log('  (0=CONNECTING, 1=OPEN, 2=CLOSED)');
  console.log('  Reconnect attempts:', reconnectAttempts);
  console.log('  Update Count:', updateCount);
  console.log('Last Trace Data:');
  console.log('  Key:', lastKey);
  console.log('  Status:', lastStatus);
  console.log('  Timestamp:', lastTs);
  console.log('───────────────────────────────');
  return {{ eventSource, updateCount, lastKey, lastStatus, lastTs }};
}};

console.log('✓ Trace viewer initialized with Server-Sent Events (SSE)');
console.log('  Mode: Event-driven (no polling!)');
console.log('  Commands: reconnectSSE(), manualRefresh(), checkStatus()');
console.log('  Open browser console to see update logs');

</script>
</body>
</html>"""

OUT_HTML.write_text(html_doc, encoding="utf-8")

print("Running from:", Path.cwd())
print("Generated:", OUT_SVG, OUT_HTML)

