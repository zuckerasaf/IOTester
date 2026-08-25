"""
Utility functions for reading Excel files and creating connector/pin objects.

This module supports two main workflows:
- `copy_and_transform_excel`: apply IOTester-specific transforms to a single sheet
  and export a file with standardized headers.
- `copy_and_transform_excel_Volcan`: split a Volcan source workbook into one Excel
  file per connector and apply row-level DI/DO/RTN/RED/INC_ENCY logic.
"""
from pathlib import Path
from typing import Optional, Dict
import argparse
import re
import sys

# Ensure the project root is on sys.path when the script is run directly.
# This allows the module to import package neighbors when executed as a script.
_PROJECT_ROOT = Path(__file__).resolve().parents[3]
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

import openpyxl

from hw_tester.hardware.pin import Pin, Connector, TestResult
from hw_tester.utils.config_loader import load_settings


def _column_letter_to_index(column_letter: str) -> int:
    """Convert Excel column letter (A, B, C, etc.) to 0-based index."""
    column_letter = column_letter.upper()
    result = 0
    for char in column_letter:
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1


def _load_excel_column_config(settings_path: Optional[str] = None) -> Dict[str, int]:
    """
    Load Excel column configuration from merged settings.

    The settings file may override default column letters. This helper resolves the
    configured letters into zero-based indices for use when reading rows.

    Returns:
        Dictionary mapping property names to 0-based column indices
    """
    if settings_path is None:
        settings = load_settings()
    else:
        settings = load_settings(path=settings_path, comm_path="Comm_settings.yaml")

    excel_cols = settings.get('ExcelColumns', {})
    
    # return {
    #     'Id': 7,                    # H
    #     'Connect': 1,               # A
    #     'Discrete_Name': 2,         # B
    #     'Signal_Name': 3,           # C
    #     'Plug': 9,                  # J
    #     'Type': 3,                  # D
    #     'Pin': 10,                  # K
    #     'Power_Expected': 13,       # N
    #     'Power_Input': 14,          # O
    #     'PullUp_Expected': 15,      # P
    #     'PullUp_Input': 16,         # Q
    #     'Logic_Pin_Input': 17,      # R
    #     'Logic_Command': 18,        # S
    #     'Logic_Expected': 19,       # T
    #     'Test_Result': 19           # T
    # }
    
    return {
        'Id': _column_letter_to_index(excel_cols.get('ID', 'H')),
        'Connect': _column_letter_to_index(excel_cols.get('Connect', 'A')),
        'Discrete_Name': _column_letter_to_index(excel_cols.get('Discrete_Name', 'B')),
        'Signal_Name': _column_letter_to_index(excel_cols.get('Signal_Name', 'C')),
        'Plug': _column_letter_to_index(excel_cols.get('Plug', 'J')),
        'Type': _column_letter_to_index(excel_cols.get('Type', 'D')),
        'Pin': _column_letter_to_index(excel_cols.get('Pin', 'K')),
        'Power_Expected': _column_letter_to_index(excel_cols.get('Power_Expected', 'N')),
        'Power_Input': _column_letter_to_index(excel_cols.get('Power_Input', 'O')),
        'PullUp_Expected': _column_letter_to_index(excel_cols.get('PullUp_Expected', 'P')),
        'PullUp_Input': _column_letter_to_index(excel_cols.get('PullUp_Input', 'Q')),
        'Logic_Pin_Input': _column_letter_to_index(excel_cols.get('Logic_Pin_Input', 'R')),
        'Logic_Command': _column_letter_to_index(excel_cols.get('Logic_Command', 'S')),
        'Logic_Expected': _column_letter_to_index(excel_cols.get('Logic_Expected', 'T')),
        'Test_Result': _column_letter_to_index(excel_cols.get('Test_Result', 'T'))
    }


def _extract_first_numeric_part(value: str) -> str:
    if not value:
        return ""
    match = re.search(r"(\d+)", value)
    return match.group(1) if match else ""


def _extract_di_do_token(value: str) -> str:
    if not value:
        return ""
    match = re.search(r"\b(DI|DO)\s*_?\s*(\d+)\b", value, re.IGNORECASE)
    if not match:
        return ""
    return f"{match.group(1).upper()}{match.group(2)}"


def _extract_voltage_from_text(value: str) -> float:
    if not value:
        return 0.0
    match = re.search(r"(\d+(?:\.\d+)?)\s*[Vv]", value)
    if match:
        return float(match.group(1))
    match = re.search(r"(\d+(?:\.\d+)?)", value)
    return float(match.group(1)) if match else 0.0


def _extract_ie_token(col_c_value: str) -> str:
    if not col_c_value:
        return "IE1"
    value = col_c_value.upper()
    if "P5" in value:
        return "IE2"
    if "P1" in value:
        return "IE1"
    return "IE1"


def _extract_pin_number(value: str) -> str:
    """Extract PIN number from text like 'PIN 2' or return 'N/A' if not found."""
    if not value:
        return 0
    # Match "PIN " followed by a number
    match = re.search(r"PIN\s+(\d+)", value, re.IGNORECASE)
    if match:
        return int(match.group(1))
    return 0


def _extract_ai_token(value: str) -> str:
    """Extract AI token from text like 'AI_5' and return 'AI5' (no underscore)."""
    if not value:
        return ""
    # Match AI followed by optional underscore and a number
    match = re.search(r"AI_?(\d+)", value, re.IGNORECASE)
    if match:
        return f"AI{match.group(1)}"
    return ""


# ============================================================================
# EXCEL TRANSFORMATION COLUMN CONFIGURATION
# ============================================================================
# Configure which source columns to read patterns from for transformation.
# Adjust these based on your Excel file structure:
#
# J1 Format (DB_Gunner_MK4):
#   Column 3: Destination 2 (signal name like "1P2 27")
#   Column 5: Color (like "Black", "White")
#   Column 6: Associated Potential/Substance (like "DI_30", "RTN_7", "24V I/O RTN_7")
#
# J31 Format (DB_Gunner_MK3):
#   Column 3: Destination 1 (signal name)
#   Column 6: Color (like "Black", "White")
#   Column 7: Associated Potential/Substance (like "DI_33", "RTN_1", "24V I/O RTN_1")
#
# Target output columns (always the same):
#   Column 14 (N): Power_Expected
#   Column 15 (O): Power_Input
#   Column 16 (P): PullUp_Expected
#   Column 17 (Q): PullUp_Input
#   Column 18 (R): logic_input
#   Column 19 (S): Logic_Command
#   Column 20 (T): Logic_Expected
#   Column 21 (U): Test_Result

EXCEL_TRANSFORM_CONFIG = {
    'DB_Gunner_MK4': {
        'signal_name_col': 3,      # Column C - Signal name (for extracting connector number)
        'color_col': 5,            # Column E - Color (for detecting "red" wires)
        'type_col': 6,             # Column F - Type/Associated Potential (DI_, DO_, RTN_, etc.)
    },
    'DB_Gunner_MK3': {
        'signal_name_col': 3,      # Column C - Signal name
        'color_col': 6,            # Column F - Color
        'type_col': 7,             # Column G - Associated Potential
        'card_col': 1,             # Column A- Card number (for INC_ENCY logic)



    },
}


def _rearrange_columns_for_tester(sheet) -> None:
    """
    Rearrange Excel columns to match the tester app's expected format.
    
    Column mapping (source -> destination):
    A -> F
    B -> E
    C -> C (no change)
    D -> J
    E -> M
    F -> L
    G -> B
    H -> I
    
    All other columns (I onwards, including transformed N-U) stay in their positions.
    This modifies the sheet in place, creating a new column arrangement.
    """
    # Read all data first (from row 1 onwards)
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    # Store only columns A-H (1-8) for remapping
    old_columns_1_to_8 = []
    for row in range(1, max_row + 1):
        row_data = []
        for col in range(1, min(9, max_col + 1)):  # Only read columns 1-8
            row_data.append(sheet.cell(row=row, column=col).value)
        old_columns_1_to_8.append(row_data)
    
    # Column mapping: old_col_index -> new_col_index (1-based)
    column_mapping = {
        1: 6,   # A -> F
        2: 5,   # B -> E
        3: 3,   # C -> C (no change)
        4: 10,  # D -> J
        5: 13,  # E -> M
        6: 12,  # F -> L
        7: 2,   # G -> B
        8: 9,   # H -> I
    }
    
    # Clear only columns 1-8 first
    for row in range(1, max_row + 1):
        for col in range(1, 9):
            sheet.cell(row=row, column=col).value = None
    
    # Write columns 1-8 to their new positions
    for row_idx, row_data in enumerate(old_columns_1_to_8, start=1):
        for old_col_idx, value in enumerate(row_data, start=1):
            new_col = column_mapping[old_col_idx]
            sheet.cell(row=row_idx, column=new_col).value = value


def copy_and_transform_excel(
    source_file: str, 
    sheet_name: Optional[str] = None,
    config_name: str = 'default',
    rearrange_columns: bool = True  # Kept for backward compatibility but ignored
) -> Path:
    """Copy an Excel file and apply IOTester column transforms.
    
    Creates a new Excel file with columns rearranged and data transformed according
    to the tester app's expected format.

    The new file is saved in the same folder with the same filename prefixed by
    ``IOTester``.
    
    Args:
        source_file: Path to the source Excel file
        sheet_name: Name of the sheet to process (None = first sheet)
        config_name: Column configuration to use ('DB_Gunner_MK4' or 'DB_Gunner_MK3')
        rearrange_columns: Deprecated, kept for compatibility (always creates rearranged file)
    
    Column mapping (source -> destination):
        Source A -> Destination F, G, H
        Source B -> Destination E
        Source C -> Destination C
        Source D -> Destination J
        Source E -> Destination M
        Source F -> Destination L
        Source G -> Destination B
        Source H -> Destination I
    """
    # Resolve relative files from the project root
    source_path = Path(source_file)
    if not source_path.is_absolute():
        source_path = Path(__file__).resolve().parents[3] / source_path

    if not source_path.exists():
        raise FileNotFoundError(f"Source Excel file not found: {source_path}")

    destination_path = source_path.with_name(f"IOTester{source_path.name}")

    # Load the source workbook
    source_workbook = openpyxl.load_workbook(source_path)
    source_sheet = source_workbook[sheet_name] if sheet_name else source_workbook.active

    # Get the column configuration for this file format
    if config_name not in EXCEL_TRANSFORM_CONFIG:
        raise ValueError(f"Unknown config_name: {config_name}. Available: {list(EXCEL_TRANSFORM_CONFIG.keys())}")
    
    config = EXCEL_TRANSFORM_CONFIG[config_name]
    
    # Create new workbook with proper structure
    dest_workbook = openpyxl.Workbook()
    dest_sheet = dest_workbook.active
    
    # Set up headers in the destination file
    headers = {
        1: 'Card',                      # A
        2: 'DISCRETE NAME',             # B
        3: 'SIGNAL NAME \ Data',        # C
        4: 'TYPE',                      # D
        5: 'CONNECTOR',                 # E
        6: 'SPS \ Vendor',              # F
        7: 'orignal',                   # G
        8: 'Pin Out Final',             # H
        9: 'FROM',                      # I
        10: 'PLUG',                     # J
        11: 'Pin',                      # K
        12: 'Color',                    # L
        13: 'AWG',                      # M
        14: 'Power_Expected',           # N
        15: 'Power_Input',              # O
        16: 'PullUp_Expected',          # P
        17: 'PullUp_Input',             # Q
        18: 'logic_input',              # R
        19: 'Logic_Command',            # S
        20: 'Logic_Expected',           # T
        21: 'Test_Result'               # U
    }
    
    # Write headers
    for col, header in headers.items():
        dest_sheet.cell(row=1, column=col, value=header)
    
    # Column mapping: source_col -> dest_col
    # Source A(1) -> Dest F(6), G(7), H(8)
    # Source B(2) -> Dest E(5)
    # Source C(3) -> Dest C(3)
    # Source D(4) -> Dest J(10)
    # Source E(5) -> Dest M(13)
    # Source F(6) -> Dest L(12)
    # Source G(7) -> Dest B(2)
    # Source H(8) -> Dest I(9)
    column_mapping = {
        #1: [6, 7, 8],  # A -> F, G, H (copy to all three)
        2: [5],        # B -> E
        3: [3],        # C -> C
        4: [10],       # D -> J
        5: [13],       # E -> M
        6: [12],       # F -> L
        7: [2],        # G -> B
        8: [9]         # H -> I
    }
    
    # Process each row (starting from row 2, skipping header)
    for src_row in range(2, source_sheet.max_row + 1):
        dest_row = src_row  # Keep same row numbers
        
        # Copy data from source to destination with column mapping
        for src_col, dest_cols in column_mapping.items():
            value = source_sheet.cell(row=src_row, column=src_col).value
            for dest_col in dest_cols:
                dest_sheet.cell(row=dest_row, column=dest_col, value=value)
        
        # Read values needed for transformation logic (from SOURCE positions)
        cell_signal = str(source_sheet.cell(row=src_row, column=config['signal_name_col']).value or "").strip()
        cell_type = str(source_sheet.cell(row=src_row, column=config['type_col']).value or "").strip()
        cell_color = str(source_sheet.cell(row=src_row, column=config['color_col']).value or "").strip()
        cell_plug = str(source_sheet.cell(row=src_row, column=4).value or "").strip()  # Source column D (becomes dest J)
        
        lower_type = cell_type.lower()
        lower_color = cell_color.lower()
        
        # Extract card numbers from different columns based on rules
        first_num_from_signal = _extract_first_numeric_part(cell_signal)  # From column C
        first_num_from_plug = _extract_first_numeric_part(cell_plug)      # From column D
        
        # Apply transformation logic - write to DESTINATION columns (N-U)
        # Rule 1: DI (Digital Input) - card number from column D
        if "di" in lower_type and "do" not in lower_type and "rtn" not in lower_type:
            dest_sheet.cell(row=dest_row, column=14, value=3.5)  # N: Power_Expected
            dest_sheet.cell(row=dest_row, column=18, value="TBD")  # R: logic_input
            di_token = _extract_di_do_token(cell_type)
            if di_token:
                dest_sheet.cell(row=dest_row, column=20, value=f"C{first_num_from_plug}_{di_token}_0")  # T: Logic_Expected
        
        # Rule 2: DO (Digital Output) - card number from column D
        if "do" in lower_type and "di" not in lower_type and "rtn" not in lower_type:
            dest_sheet.cell(row=dest_row, column=14, value=0)  # N: Power_Expected
            dest_sheet.cell(row=dest_row, column=16, value=4)  # P: PullUp_Expected
            do_token = _extract_di_do_token(cell_type)
            if do_token:
                dest_sheet.cell(row=dest_row, column=17, value=f"C{first_num_from_plug}_{do_token}_1")  # Q: PullUp_Input
        
        # Rule 3: RTN (Return)
        if "rtn" in lower_type:
            dest_sheet.cell(row=dest_row, column=14, value=0)  # N: Power_Expected
            dest_sheet.cell(row=dest_row, column=16, value=0)  # P: PullUp_Expected
            dest_sheet.cell(row=dest_row, column=17, value="G")  # Q: PullUp_Input
        
        # Rule 4: Red wires (Power)
        if "red" in lower_color:
            voltage = _extract_voltage_from_text(cell_type)
            dest_sheet.cell(row=dest_row, column=14, value=voltage)  # N: Power_Expected
            dest_sheet.cell(row=dest_row, column=15, value="P")  # O: Power_Input
        
        # Rule 5: INC_ENCY_A - card number from column D
        if "inc_ency_a" in lower_type:
            dest_sheet.cell(row=dest_row, column=14, value=3.5)  # N: Power_Expected
            dest_sheet.cell(row=dest_row, column=18, value="TBD")  # R: logic_input
            ie_token = _extract_ie_token(cell_signal)
            dest_sheet.cell(row=dest_row, column=20, value=f"C{first_num_from_plug}_{ie_token}_1")  # T: Logic_Expected
        
        # Rule 6: INC_ENCY_B - card number from column D
        if "inc_ency_b" in lower_type:
            dest_sheet.cell(row=dest_row, column=14, value=3.5)  # N: Power_Expected
            dest_sheet.cell(row=dest_row, column=18, value="TBD")  # R: logic_input
            ie_token = _extract_ie_token(cell_signal)
            dest_sheet.cell(row=dest_row, column=20, value=f"C{first_num_from_plug}_{ie_token}_65535")  # T: Logic_Expected
        
        # Rule 7-9: Card number from source column D (PLUG) - check for 1P, 2P, 3P patterns
        if cell_plug:
            if "1P" in cell_plug.upper():
                dest_sheet.cell(row=dest_row, column=1, value=1)  # A: Card
            elif "2P" in cell_plug.upper():
                dest_sheet.cell(row=dest_row, column=1, value=2)  # A: Card
            elif "3P" in cell_plug.upper():
                dest_sheet.cell(row=dest_row, column=1, value=3)  # A: Card
        
        # Rule 10: Extract PIN number from column C (SIGNAL NAME) and put in F, G, H
        pin_number_value = _extract_pin_number(cell_signal)
        dest_sheet.cell(row=dest_row, column=6, value=pin_number_value)  # F: SPS \ Vendor
        dest_sheet.cell(row=dest_row, column=7, value=pin_number_value)  # G: orignal
        dest_sheet.cell(row=dest_row, column=8, value=pin_number_value)  # H: Pin Out Final
        
        # Rule 11: AI token handling - if TYPE contains AI (e.g., AI_5)
        ai_token = _extract_ai_token(cell_type)
        if ai_token:
            dest_sheet.cell(row=dest_row, column=14, value=0)  # N: Power_Expected
            dest_sheet.cell(row=dest_row, column=18, value="TBD")  # R: logic_input
            # Build Logic_Expected: C{card_num}_AI{num}_10
            first_num_from_plug = _extract_first_numeric_part(cell_plug)
            dest_sheet.cell(row=dest_row, column=20, value=f"C{first_num_from_plug}_{ai_token}_10")  # T: Logic_Expected
    
    # Save and close
    dest_workbook.save(destination_path)
    dest_workbook.close()
    source_workbook.close()
    
    return destination_path






def _extract_card_value_from_sheet_name(sheet_name: str) -> str:
    """Extract a numeric card value from a sheet title like 'Spider Card #1'."""
    if not sheet_name:
        return ""
    match = re.search(r"#\s*(\d+)", sheet_name)
    if match:
        return match.group(1)
    match = re.search(r"(\d+)", sheet_name)
    return match.group(1) if match else ""


def _is_volcan_connector(connector_value: str) -> bool:
    """Return True for valid Volcan connector values J1..J26."""
    if not connector_value:
        return False
    normalized = str(connector_value).strip().upper()
    return bool(re.fullmatch(r"J([1-9]|1\d|2[0-6])", normalized))


def copy_and_transform_excel_Volcan(source_file: str, sheet_name: Optional[str] = None) -> Dict[str, Path]:
    """Generate Volcan connector workbooks from a source Excel file."""
    source_path = Path(source_file)
    if not source_path.is_absolute():
        source_path = Path(__file__).resolve().parents[3] / source_path

    if not source_path.exists():
        raise FileNotFoundError(f"Source Excel file not found: {source_path}")

    workbook = openpyxl.load_workbook(source_path)
    sheets = [workbook[sheet_name]] if sheet_name else workbook.worksheets

    # Group rows by connector, preserving the source row and the card number
    # extracted from the worksheet title.
    rows_by_connector = {}
    for sheet in sheets:
        card_value = _extract_card_value_from_sheet_name(sheet.title)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            connector_value = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
            if not _is_volcan_connector(connector_value):
                continue
            connector_name = connector_value.upper()
            rows_by_connector.setdefault(connector_name, []).append((row, card_value))

    result_paths = {}
    for connector_name, connector_rows in rows_by_connector.items():
        target_workbook = openpyxl.Workbook()
        target_sheet = target_workbook.active
        headers = [
            "Card",
            "Discrete Mame",
            "Signal Name",
            "Type",
            "Connector",
            "SPS",
            "Orignal",
            "Pin Out",
            "From",
            "Plug",
            "Pin",
            "Color",
            "AWG",
            "Power_Expected",
            "Power_Input",
            "PullUp_Expected",
            "PullUp_Input",
            "logic_input",
            "Logic_Command",
            "Logic_Expected",
            "Test_Result",
        ]
        target_sheet.append(headers)

        for row, card_value in connector_rows:
            # Build the target row template from the source row values.
            # Columns F and G are intentionally blank in the target file.
            target_row = [
                card_value,
                row[1] if len(row) > 1 else None,
                row[2] if len(row) > 2 else None,
                row[3] if len(row) > 3 else None,
                row[4] if len(row) > 4 else None,
                None,
                None,
                row[5] if len(row) > 5 else None,
                row[6] if len(row) > 6 else None,
                row[7] if len(row) > 7 else None,
                row[8] if len(row) > 8 else None,
                row[9] if len(row) > 9 else None,
                row[10] if len(row) > 10 else None,
                row[13] if len(row) > 13 else None,
                row[14] if len(row) > 14 else None,
                row[15] if len(row) > 15 else None,
                row[16] if len(row) > 16 else None,
                row[17] if len(row) > 17 else None,
                row[18] if len(row) > 18 else None,
                row[19] if len(row) > 19 else None,
                row[20] if len(row) > 20 else None,
            ]

            # Normalize the important source fields used by the update rules.
            source_b = str(target_row[1] or "").strip()
            source_c = str(target_row[2] or "").strip()
            source_i = str(target_row[8] or "").strip()
            source_j = str(target_row[9] or "").strip()
            source_l = str(target_row[11] or "").strip()
            card_number = str(card_value).strip()

            di_match = re.search(r"\bDI[_ ]?(\d+)\b", source_b, re.IGNORECASE)
            if di_match:
                di_token = f"DI{di_match.group(1)}"
                target_row[13] = 3.5
                target_row[17] = "TBD"
                target_row[19] = f"C{card_number}_{di_token}_0"

            do_match = re.search(r"\bDO[_ ]?(\d+)\b", source_b, re.IGNORECASE)
            if do_match:
                do_token = f"DO{do_match.group(1)}"
                target_row[13] = 0
                target_row[15] = 4
                target_row[16] = f"C{card_number}_{do_token}_1"

            if "RTN" in source_j.upper():
                # Return rows are wired to ground.
                target_row[13] = 0
                target_row[15] = 0
                target_row[16] = "G"

            if "RED" in source_l.upper() and "TERMINAL BLOCK" in source_i.upper():
                # Terminal block red wires are treated as power lines.
                target_row[13] = _extract_voltage_from_text(source_j)
                target_row[14] = "P"

            if "INC_ENCY" in source_c.upper():
                # INC_ENCY logic uses a special IE token derived from the source label.
                ie_token = "IE2" if "P5" in source_j.upper() else "IE1"
                suffix = "1" if "A+" in source_c.upper() else "65535"
                target_row[13] = 3.5
                target_row[17] = "TBD"
                target_row[19] = f"C{card_number}_{ie_token}_{suffix}"

            target_sheet.append(target_row)

        destination_path = source_path.with_name(f"Volcan_{connector_name}.xlsx")
        target_workbook.save(destination_path)
        target_workbook.close()
        result_paths[connector_name] = destination_path

    workbook.close()
    return result_paths


def load_connector_from_excel(
    file_name: str = "J17_Armant.xlsx",
    db_path: str = "tests/DB",
    connector_id: str = "J17",
    sheet_name: Optional[str] = None,
    settings_path: Optional[str] = None
) -> Connector:
    """
    Read an Excel file and create a Connector with pins.

    This helper is used by the IOTester application to load connector pin
    definitions from a workbook into the internal Connector/Pin model.

    Args:
        file_name: Name of the Excel file (default: "J17_Armant.xlsx")
        db_path: Path to the database folder (default: "tests/DB")
        connector_id: ID for the connector (default: "J17")
        sheet_name: Name of the sheet to read (default: first sheet)
        settings_path: Path to settings.yaml (default: auto-detected)

    Returns:
        Connector object populated with pins from the Excel file

    Raises:
        FileNotFoundError: If the Excel file doesn't exist
        ValueError: If the Excel file format is invalid

    Excel columns are configured in settings.yaml under ExcelColumns section.
    Data starts from row 2 (row 1 is assumed to be header).
    """
    # Load column configuration from settings and convert column letters to indices.
    col_map = _load_excel_column_config(settings_path)
    
    # Build the full path to the Excel file.
    # If db_path is absolute, use it directly; otherwise, resolve it from the project root.
    db_path_obj = Path(db_path)
    if db_path_obj.is_absolute():
        excel_path = db_path_obj / file_name
    else:
        project_root = Path(__file__).resolve().parents[3]
        excel_path = project_root / db_path / file_name
    
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    # Load the workbook and get the sheet
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    
    # Create the connector
    connector = Connector(id=connector_id)
    
    # Read data starting from row 2 (skip header)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Read values from configured columns
        pin_id = row[col_map['Id']] if len(row) > col_map['Id'] else None
        connect = row[col_map['Connect']] if len(row) > col_map['Connect'] else None
        discrete_name = row[col_map['Discrete_Name']] if len(row) > col_map['Discrete_Name'] else None
        signal_name = row[col_map['Signal_Name']] if len(row) > col_map['Signal_Name'] else None
        plug = row[col_map['Plug']] if len(row) > col_map['Plug'] else None
        type_value = row[col_map['Type']] if len(row) > col_map['Type'] else None
        pin_number = row[col_map['Pin']] if len(row) > col_map['Pin'] else None
        power_expected = row[col_map['Power_Expected']] if len(row) > col_map['Power_Expected'] else None
        power_input = row[col_map['Power_Input']] if len(row) > col_map['Power_Input'] else None
        pullup_expected = row[col_map['PullUp_Expected']] if len(row) > col_map['PullUp_Expected'] else None
        pullup_input = row[col_map['PullUp_Input']] if len(row) > col_map['PullUp_Input'] else None
        logic_pin_input = row[col_map['Logic_Pin_Input']] if len(row) > col_map['Logic_Pin_Input'] else None
        logic_command = row[col_map['Logic_Command']] if len(row) > col_map['Logic_Command'] else None
        logic_expected = row[col_map['Logic_Expected']] if len(row) > col_map['Logic_Expected'] else None
        
        # Skip empty rows or rows without pin ID
        if not pin_id:
            continue
        
        # Convert to string and strip whitespace
        pin_id = str(pin_id).strip()
        connect = str(connect).strip() if connect else ""
        discrete_name_value = str(discrete_name).strip() if discrete_name else ""
        signal_name_value = str(signal_name).strip() if signal_name else ""
        plug_value = str(plug).strip() if plug else ""
        type_value_str = str(type_value).strip() if type_value else ""
        pin_number_value = str(pin_number).strip() if pin_number else ""
        
        # Parse power input - keep as string from Excel
        power_input_value = str(power_input).strip() if power_input else ""
        
        # Parse power expected voltage
        try:
            power_expected_value = float(power_expected) if power_expected else 0.0
        except (ValueError, TypeError):
            power_expected_value = 0.0
        
        # Parse pullup expected voltage
        try:
            pullup_expected_value = float(pullup_expected) if pullup_expected else 0.0
        except (ValueError, TypeError):
            pullup_expected_value = 0.0
        
        # Parse pullup input - keep as string from Excel
        pullup_input_value = str(pullup_input).strip() if pullup_input else ""
        
        # Parse logic pin input - keep as string from Excel
        logic_pin_input_value = str(logic_pin_input).strip() if logic_pin_input else ""
        # Parse logic command - keep as string from Excel
        logic_command_value = str(logic_command).strip() if logic_command else ""

        # Parse logic expected - keep as string from Excel
        logic_expected_value = str(logic_expected).strip() if logic_expected else ""
        
        # Create pin with new structure
        pin = Pin(
            Id=pin_id,
            Connect=connect,
            Discrete_Name=discrete_name_value,
            Signal_Name=signal_name_value,
            Plug=plug_value,
            Type=type_value_str,
            Pin=pin_number_value,
            Power_Expected=power_expected_value,
            Power_Measured=0.0,
            Power_Result=TestResult.NO_RESULT,
            PullUp_Expected=pullup_expected_value,
            PullUp_Measured=0.0,
            PullUp_Result=TestResult.NO_RESULT,
            Power_Input=power_input_value,
            PullUp_Input=pullup_input_value,
            Logic_Pin_Input=logic_pin_input_value,
            Logic_Command=logic_command_value,
            Logic_Expected=logic_expected_value,
            Logic_DI_Result=TestResult.NO_RESULT
        )
        
        connector.add_pin(pin)
    
    workbook.close()
    
    return connector



def main() -> None:
    """Process all Excel files in a specified folder."""
    
    # Specify the folder containing Excel files to transform
    source_folder = Path(r"C:\ArduinoProject\IO_Tester\DB_Gunner_MK3")
    
    # Optional: specify sheet name (None = use first sheet automatically)
    sheet_name = None  # Changed from "Sheet1" to None to handle different sheet names
    
    # Specify column configuration:
    config_name = 'DB_Gunner_MK3'
    
    # Rearrange columns to match tester app format (True = enabled)
    # This remaps columns: A->F, B->E, C->C, D->J, E->M, F->L, G->B, H->I
    rearrange_columns = True
    
    # Find all Excel files in the folder
    excel_files = list(source_folder.glob("*.xlsx")) + list(source_folder.glob("*.xls"))
    
    # Filter out files that already start with "IOTester" to avoid reprocessing
    excel_files = [f for f in excel_files if not f.name.startswith("IOTester")]
    
    if not excel_files:
        print(f"No Excel files found in: {source_folder}")
        return
    
    print(f"Found {len(excel_files)} Excel file(s) to process in: {source_folder}")
    print(f"Using column configuration: '{config_name}'")
    print(f"Column rearrangement: {'Enabled' if rearrange_columns else 'Disabled'}")
    print("-" * 80)
    
    # Process each Excel file
    success_count = 0
    failed_count = 0
    
    for excel_file in excel_files:
        try:
            print(f"\nProcessing: {excel_file.name}")
            destination = copy_and_transform_excel(
                source_file=str(excel_file),
                sheet_name=sheet_name,
                config_name=config_name,
                rearrange_columns=rearrange_columns
            )
            print(f"✓ Created: {destination.name}")
            success_count += 1
        except Exception as e:
            print(f"✗ Failed to process {excel_file.name}: {e}")
            failed_count += 1
    
    # Summary
    print("\n" + "=" * 80)
    print(f"Processing complete!")
    print(f"  Success: {success_count} file(s)")
    print(f"  Failed:  {failed_count} file(s)")
    print("=" * 80)
    
    # Uncomment below to process Volcan files
    # copy_and_transform_excel_Volcan(
    #     source_file=r"C:\ArduinoProject\IO_Tester\DB_Volcan\HSID Volcan V8 REV03.xlsx"
    # )


if __name__ == "__main__":
    main()