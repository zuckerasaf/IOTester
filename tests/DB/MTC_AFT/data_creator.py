from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import re


SOURCE_XLSX = Path(r"HSID AFT V_7 15.05.2025.xlsx")
TARGET_TEMPLATE_XLSX = Path(r"J_structure.xlsx")  # used only as structure/templateJ12

SHEET_PREFIXES = ("CARD",)


def find_header_row(df_raw: pd.DataFrame, header_key: str = "DISCRETE NAME") -> int:
    tmp = df_raw.astype(str).applymap(lambda x: x.strip())
    matches = (tmp == header_key).any(axis=1)
    idx_list = tmp.index[matches].tolist()
    if not idx_list:
        raise ValueError(f"Could not find header row containing '{header_key}'.")
    return int(idx_list[0])


def read_card_sheet(source_path: Path, sheet_name: str) -> pd.DataFrame:
    raw = pd.read_excel(source_path, sheet_name=sheet_name, header=None, dtype=str)
    header_row = find_header_row(raw)

    df = pd.read_excel(source_path, sheet_name=sheet_name, header=header_row, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how="all")

    return df


def extract_connector_rows(source_path: Path, connector_value: str) -> pd.DataFrame:
    xl = pd.ExcelFile(source_path)
    card_sheets = [s for s in xl.sheet_names if s.upper().startswith(SHEET_PREFIXES)]

    rows = []
    for sheet in card_sheets:
        df = read_card_sheet(source_path, sheet)

        if "CONNECTOR" not in df.columns:
            continue

        mask = df["CONNECTOR"].astype(str).str.strip() == connector_value
        df = df.loc[mask].copy()

        if not df.empty:            # Extract card number from sheet name (e.g., "CARD2" -> "2")
            card_num = sheet.upper().replace("CARD", "").strip()
            df.insert(0, "CARD", card_num)  # Add as first column (column A)            df["SOURCE_SHEET"] = sheet  # optional traJ17ceability
            rows.append(df)

    if not rows:
        return pd.DataFrame()

    return pd.concat(rows, ignore_index=True)


def apply_autofill_rules(df: pd.DataFrame) -> pd.DataFrame:
    """Apply autofill rules to populate empty cells based on other column values."""
    if df.empty:
        return df
    
    df = df.copy()
    
    # Ensure required columns exist
    required_cols = ['CARD', 'DISCRETE NAME', 'PLUG', 'Pin', 
                     'Power_Expected', 'Power_Input', 'PullUp_Expected', 
                     'PullUp_Input', 'logic_input', 'Logic_Expected']
    for col in required_cols:
        if col not in df.columns:
            df[col] = None
    
    for idx, row in df.iterrows():
        discrete_name = str(row.get('DISCRETE NAME', '')).strip().upper()
        plug = str(row.get('PLUG', '')).strip().upper()
        pin = str(row.get('Pin', '')).strip().upper()
        card_num = str(row.get('CARD', '')).strip()
        
        # Priority 1: Check DISCRETE NAME first (Rules 1, 2, 8)
        # Rule 1: If B (DISCRETE NAME) contains "DO"
        if 'DO_' in discrete_name or discrete_name.startswith('DO'):
            df.at[idx, 'Power_Expected'] = 0
            df.at[idx, 'PullUp_Expected'] = 21
            # Extract just the DO part (e.g., "DO_16" -> "DO16")
            discrete_clean = discrete_name.replace('_', '')
            df.at[idx, 'PullUp_Input'] = f"C{card_num}_{discrete_clean}_1"
            continue  # Skip PLUG checks for this row
        
        # Rule 2: If B (DISCRETE NAME) contains "DI"
        if 'DI_' in discrete_name or discrete_name.startswith('DI'):
            df.at[idx, 'Power_Expected'] = 3.5
            df.at[idx, 'logic_input'] = "?"
            discrete_clean = discrete_name.replace('_', '')
            df.at[idx, 'Logic_Expected'] = f"C{card_num}_{discrete_clean}_0"
            continue  # Skip PLUG checks for this row
        
        # Rule 8: If B (DISCRETE NAME) contains "AI"
        if 'AI_' in discrete_name or discrete_name.startswith('AI'):
            df.at[idx, 'Power_Expected'] = 0
            df.at[idx, 'logic_input'] = "?"
            discrete_clean = discrete_name.replace('_', '')
            df.at[idx, 'Logic_Expected'] = f"C{card_num}_{discrete_clean}_10"
            continue  # Skip PLUG checks for this row
        
        # Priority 2: Check PLUG and Pin columns (Rules 3-7)
        # Rule 6: If J contains "I\O RTN" (check before Rule 7 to avoid conflict)
        if 'I\\O RTN' in plug or 'I/O RTN' in plug:
            df.at[idx, 'Power_Expected'] = 0
            df.at[idx, 'PullUp_Expected'] = 0
            df.at[idx, 'PullUp_Input'] = "G"
        
        # Rule 7: If J contains "I\O" (but not "I\O RTN")
        elif ('I\\O' in plug or 'I/O' in plug) and 'RTN' not in plug:
            # Extract voltage number (e.g., "24V I\O" -> 24)
            voltage_match = re.search(r'(\d+)V', plug)
            if voltage_match:
                voltage = int(voltage_match.group(1))
                df.at[idx, 'Power_Expected'] = voltage
                df.at[idx, 'Power_Input'] = "P"
        
        # Rule 3: If J contains "0-28V" and K contains "+"
        elif '0-28V' in plug and '+' in pin:
            df.at[idx, 'Power_Expected'] = 24
            df.at[idx, 'Power_Input'] = "C1_AO5_10"
        
        # Rule 4: If J contains "0-5V" and K contains "+"
        elif '0-5V' in plug and '+' in pin:
            df.at[idx, 'Power_Expected'] = 5
            df.at[idx, 'Power_Input'] = "C1_AO2_10"
        
        # Rule 5: If J contains "0-5V" and K contains "-"
        elif '0-5V' in plug and '-' in pin:
            df.at[idx, 'Power_Expected'] = 0
            df.at[idx, 'PullUp_Expected'] = 0
            df.at[idx, 'PullUp_Input'] = "G"
    
    return df


def write_into_template(template_path: Path, output_path: Path, data: pd.DataFrame) -> None:
    wb = load_workbook(template_path)
    ws = wb.worksheets[0]

    template_headers = [
        str(ws.cell(row=1, column=c).value).strip()
        if ws.cell(row=1, column=c).value else ""
        for c in range(1, ws.max_column + 1)
    ]

    # Clear old data
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    if data.empty:
        wb.save(output_path)
        return

    data = data.copy()
    data.columns = [str(c).strip() for c in data.columns]

    for r_idx, (_, row) in enumerate(data.iterrows(), start=2):
        # Always write CARD column to column A if it exists
        if "CARD" in data.columns:
            ws.cell(row=r_idx, column=1).value = row.get("CARD")
        
        for c_idx, header in enumerate(template_headers, start=1):
            if header and header in data.columns:
                ws.cell(row=r_idx, column=c_idx).value = row.get(header)

    wb.save(output_path)


def main() -> None:
    connector_value = input("Enter CONNECTOR value (e.g. J5): ").strip()

    if not connector_value:
        raise ValueError("CONNECTOR value cannot be empty")

    output_file = Path(f"{connector_value}.xlsx")

    df = extract_connector_rows(SOURCE_XLSX, connector_value)
    
    # Apply autofill rules to populate empty cells
    df = apply_autofill_rules(df)
    
    write_into_template(TARGET_TEMPLATE_XLSX, output_file, df)

    print(f"✔ Done: {len(df)} rows written to '{output_file.name}'")


if __name__ == "__main__":
    main()
